from __future__ import annotations

import hashlib
from pathlib import Path

from openpyxl import load_workbook

from app.schemas import (
    BindingStrategy,
    FileType,
    TemplateBinding,
    TemplateField,
    TemplateManifest,
)
from app.services.xlsx.render import render_xlsx


def _hash(path: Path) -> str:
    return hashlib.sha256(path.read_bytes()).hexdigest()


def test_xlsx_only_changes_bound_values(xlsx_fixture: Path, tmp_path: Path) -> None:
    before_hash = _hash(xlsx_fixture)
    source = load_workbook(xlsx_fixture, data_only=False)
    source_sheet = source["Price Comparison"]
    snapshot = {
        "merges": sorted(str(value) for value in source_sheet.merged_cells.ranges),
        "widths": {key: value.width for key, value in source_sheet.column_dimensions.items()},
        "heights": {key: value.height for key, value in source_sheet.row_dimensions.items()},
        "formula": source_sheet["D8"].value,
        "print_area": str(source_sheet.print_area),
        "freeze": source_sheet.freeze_panes,
    }
    manifest = TemplateManifest(
        template_id="fixture",
        name="Comparison",
        file_type=FileType.XLSX,
        original_file="comparison.xlsx",
        fields=[
            TemplateField(
                id="project_name",
                label="ชื่อโครงการ",
                binding=TemplateBinding(
                    strategy=BindingStrategy.CELL,
                    sheet="Price Comparison",
                    cell="B3",
                ),
            ),
            TemplateField(
                id="vendor_a_price",
                label="ราคาผู้เสนอ A",
                binding=TemplateBinding(
                    strategy=BindingStrategy.CELL,
                    sheet="Price Comparison",
                    cell="D10",
                ),
            ),
        ],
    )
    output = tmp_path / "rendered.xlsx"
    warnings = render_xlsx(
        xlsx_fixture,
        output,
        manifest,
        {"project_name": "จัดซื้อเครื่องสำรองไฟฟ้า", "vendor_a_price": 43000},
    )

    assert warnings == []
    assert _hash(xlsx_fixture) == before_hash
    rendered = load_workbook(output, data_only=False)["Price Comparison"]
    assert rendered["B3"].value == "จัดซื้อเครื่องสำรองไฟฟ้า"
    assert rendered["D10"].value == 43000
    assert sorted(str(value) for value in rendered.merged_cells.ranges) == snapshot["merges"]
    assert {key: value.width for key, value in rendered.column_dimensions.items()} == snapshot[
        "widths"
    ]
    assert {key: value.height for key, value in rendered.row_dimensions.items()} == snapshot[
        "heights"
    ]
    assert rendered["D8"].value == snapshot["formula"]
    assert str(rendered.print_area) == snapshot["print_area"]
    assert rendered.freeze_panes == snapshot["freeze"]
