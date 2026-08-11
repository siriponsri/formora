from __future__ import annotations

from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


def _header_footer(value) -> dict[str, Any]:
    return {
        "left": value.left.text,
        "center": value.center.text,
        "right": value.right.text,
    }


def inspect_xlsx(path: str | Path) -> dict[str, Any]:
    source = Path(path)
    workbook = load_workbook(source, data_only=False, keep_links=True)
    sheets: list[dict[str, Any]] = []
    for worksheet in workbook.worksheets:
        used_cells = []
        formulas = []
        for row in worksheet.iter_rows():
            for cell in row:
                if cell.value is None:
                    continue
                cell_map = {
                    "coordinate": cell.coordinate,
                    "value": cell.value,
                    "data_type": cell.data_type,
                    "style_id": cell.style_id,
                    "number_format": cell.number_format,
                }
                used_cells.append(cell_map)
                if cell.data_type == "f":
                    formulas.append({"coordinate": cell.coordinate, "formula": cell.value})

        columns = {
            get_column_letter(index): {
                "width": worksheet.column_dimensions[get_column_letter(index)].width
            }
            for index in range(1, worksheet.max_column + 1)
            if worksheet.column_dimensions[get_column_letter(index)].width is not None
        }
        rows = {
            str(index): {"height": dimension.height}
            for index, dimension in worksheet.row_dimensions.items()
            if dimension.height is not None
        }
        data_validations = []
        if worksheet.data_validations:
            for validation in worksheet.data_validations.dataValidation:
                data_validations.append(
                    {
                        "type": validation.type,
                        "ranges": str(validation.sqref),
                        "formula1": validation.formula1,
                        "formula2": validation.formula2,
                    }
                )

        sheets.append(
            {
                "name": worksheet.title,
                "max_row": worksheet.max_row,
                "max_column": worksheet.max_column,
                "dimension": worksheet.calculate_dimension(),
                "used_cells": used_cells,
                "formulas": formulas,
                "merged_ranges": [str(value) for value in worksheet.merged_cells.ranges],
                "columns": columns,
                "rows": rows,
                "freeze_panes": str(worksheet.freeze_panes) if worksheet.freeze_panes else None,
                "print_area": str(worksheet.print_area) if worksheet.print_area else None,
                "orientation": worksheet.page_setup.orientation,
                "paper_size": worksheet.page_setup.paperSize,
                "odd_header": _header_footer(worksheet.oddHeader),
                "odd_footer": _header_footer(worksheet.oddFooter),
                "data_validations": data_validations,
            }
        )
    return {
        "file_type": "xlsx",
        "filename": source.name,
        "sheet_names": workbook.sheetnames,
        "sheets": sheets,
    }
