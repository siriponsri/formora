from __future__ import annotations

import shutil
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from app.schemas import BindingStrategy, TemplateManifest


def render_xlsx(
    original_path: str | Path,
    output_path: str | Path,
    manifest: TemplateManifest,
    content: dict[str, Any],
) -> list[str]:
    source = Path(original_path)
    target = Path(output_path)
    target.parent.mkdir(parents=True, exist_ok=True)
    shutil.copy2(source, target)
    workbook = load_workbook(target, data_only=False, keep_links=True)
    warnings: list[str] = []

    for field in manifest.fields:
        binding = field.binding
        if binding.strategy != BindingStrategy.CELL:
            continue
        if binding.sheet and binding.sheet not in workbook.sheetnames:
            warnings.append(f"Sheet '{binding.sheet}' for field '{field.label}' was not found.")
            continue
        worksheet = workbook[binding.sheet] if binding.sheet else workbook.worksheets[0]
        value = content.get(field.id)
        if value is None and field.required:
            warnings.append(f"Required field '{field.label}' had no value.")
        worksheet[binding.cell or "A1"].value = value

    workbook.save(target)
    return warnings
